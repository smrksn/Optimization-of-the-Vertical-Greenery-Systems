import pandas as pd
from openpyxl import load_workbook
import os

substrates_dry = ['1. Preliminary data process\Input\substrates_dry_v1.xlsx', '1. Preliminary data process\Input\substrates_dry_v2.xlsx']
substrates_sat = ['1. Preliminary data process\Input\substrates_sat_v1.xlsx', '1. Preliminary data process\Input\substrates_sat_v2.xlsx']
mineral_wool_dry = ['1. Preliminary data process\Input\mineral_wool_dry_v1.xlsx', '1. Preliminary data process\Input\mineral_wool_dry_v2.xlsx']
mineral_wool_sat = ['1. Preliminary data process\Input\mineral_wool_sat_v1.xlsx', '1. Preliminary data process\Input\mineral_wool_sat_v2.xlsx']

data = [substrates_dry, substrates_sat, mineral_wool_dry, mineral_wool_sat]
    

for instance in data:
    df1 = pd.read_excel(instance[0], header=[0, 1])
    df2 = pd.read_excel(instance[1], header=[0, 1])

    mean_df = (df1 + df2) / 2

    if 'substrates' in instance[0] and 'dry' in instance[0]:
        filename = 'substrates_dry.xlsx'
    elif 'substrates' in instance[0] and 'sat' in instance[0]:
        filename = 'substrates_sat.xlsx'
    elif 'mineral_wool' in instance[0] and 'dry' in instance[0]:
        filename = 'mineral_wool_dry.xlsx'
    elif 'mineral_wool' in instance[0] and 'sat' in instance[0]:
        filename = 'mineral_wool_sat.xlsx'

    folder_path = r'1. Preliminary data process\Output'
    if not os.path.exists(folder_path):
        os.makedirs(folder_path)
    writer = pd.ExcelWriter(os.path.join(folder_path, filename), engine='xlsxwriter')
    mean_df.to_excel(writer, sheet_name='Impedance Ratio')
    writer.close()

    book = load_workbook(os.path.join(folder_path, filename))
    sheet = book['Impedance Ratio']

    # Unmerge cells in row 1
    merged_ranges = list(sheet.merged_cells)
    for merged_range in merged_ranges:
        if merged_range.min_row == 1:
            sheet.unmerge_cells(str(merged_range))
    sheet.delete_cols(1)
    sheet.delete_rows(3)
    book.save(os.path.join(folder_path, filename))